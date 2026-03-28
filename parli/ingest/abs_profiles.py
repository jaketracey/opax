"""
parli.ingest.abs_profiles -- Ingest ABS Census electorate demographic profiles.

Downloads 2021 Census General Community Profile datapacks for Commonwealth
Electoral Divisions (CEDs) from the ABS website, extracts key demographic
indicators, and loads them into the electorate_demographics table.

Data sources:
  - G01: Population totals, indigenous status, country of birth
  - G02: Median age, income, rent, mortgage, household size
  - G37: Housing tenure (owned outright, mortgage, rented)
  - G43: Labour force, unemployment, education qualifications

Cross-reference analysis compares electorate demographics with how
MPs vote on related policy issues, surfacing representation gaps.

Usage:
    # Download and ingest ABS Census CED profiles
    python -m parli.ingest.abs_profiles

    # Skip download (use cached zip)
    python -m parli.ingest.abs_profiles --no-download

    # Run cross-reference analysis only
    python -m parli.ingest.abs_profiles --analyze

    # Full pipeline
    python -m parli.ingest.abs_profiles --analyze
"""

import argparse
import csv
import io
import json
import os
import sqlite3
import zipfile
from pathlib import Path

import requests

from parli.schema import get_db, init_db

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "OPAX/1.0 (parliamentary transparency research; https://opax.com.au)"
})

CACHE_DIR = Path("~/.cache/autoresearch/abs").expanduser()
ZIP_URL = "https://www.abs.gov.au/census/find-census-data/datapacks/download/2021_GCP_CED_for_AUS_short-header.zip"
ZIP_FILE = CACHE_DIR / "2021_GCP_CED_AUS.zip"

# Inside the zip, the folder has a typo: "Electroral" not "Electoral"
DATA_SUBDIR = "2021 Census GCP Commonwealth Electroral Division for AUS"
METADATA_SUBDIR = "Metadata"

# ABS CED code -> state mapping (first digit of CED number)
CED_STATE_MAP = {
    "1": "NSW",
    "2": "VIC",
    "3": "QLD",
    "4": "SA",
    "5": "WA",
    "6": "TAS",
    "7": "NT",
    "8": "ACT",
    "9": "OT",   # Other Territories
}


def download_datapack() -> Path:
    """Download the ABS 2021 Census CED datapack zip if not cached."""
    CACHE_DIR.mkdir(parents=True, exist_ok=True)

    if ZIP_FILE.exists():
        size_mb = ZIP_FILE.stat().st_size / (1024 * 1024)
        print(f"  Using cached zip: {ZIP_FILE} ({size_mb:.1f} MB)")
        return ZIP_FILE

    print(f"  Downloading ABS Census CED datapack from {ZIP_URL} ...")
    resp = SESSION.get(ZIP_URL, stream=True, timeout=120)
    resp.raise_for_status()

    with open(ZIP_FILE, "wb") as f:
        for chunk in resp.iter_content(chunk_size=8192):
            f.write(chunk)

    size_mb = ZIP_FILE.stat().st_size / (1024 * 1024)
    print(f"  Downloaded {size_mb:.1f} MB -> {ZIP_FILE}")
    return ZIP_FILE


def read_csv_from_zip(zf: zipfile.ZipFile, filename: str) -> list[dict]:
    """Read a CSV file from within the zip archive."""
    # Find the actual path inside the zip (may be nested)
    target = None
    for name in zf.namelist():
        if name.endswith(filename):
            target = name
            break

    if target is None:
        raise FileNotFoundError(f"CSV {filename} not found in zip")

    with zf.open(target) as f:
        text = io.TextIOWrapper(f, encoding="utf-8-sig")
        reader = csv.DictReader(text)
        return list(reader)


def read_geography_mapping(zf: zipfile.ZipFile) -> dict[str, str]:
    """Read the geography descriptor Excel file to map CED codes -> names.

    Falls back to parsing from the zip if openpyxl is available.
    """
    # Try to read the Excel metadata file
    geo_file = None
    for name in zf.namelist():
        if "geog_desc" in name and name.endswith(".xlsx"):
            geo_file = name
            break

    if geo_file is None:
        print("  Warning: geography descriptor file not found in zip")
        return {}

    try:
        import openpyxl

        # Extract to temp location
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(zf.read(geo_file))
            tmp_path = tmp.name

        wb = openpyxl.load_workbook(tmp_path, read_only=True)
        # Look for the Non-ABS structures sheet (contains CED mappings)
        mapping = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                if row and len(row) >= 4 and row[0] == "CED":
                    code = str(row[1])  # e.g., "CED101"
                    name = str(row[3])  # e.g., "Banks"
                    mapping[code] = name

        os.unlink(tmp_path)
        wb.close()
        return mapping

    except ImportError:
        print("  Warning: openpyxl not available, will use G01 codes without names")
        return {}


def get_state_from_code(ced_code: str) -> str:
    """Extract state from CED code (e.g., CED101 -> NSW, CED201 -> VIC)."""
    # Strip "CED" prefix, first digit is the state
    num = ced_code.replace("CED", "")
    if num and num[0] in CED_STATE_MAP:
        return CED_STATE_MAP[num[0]]
    return "Unknown"


def safe_float(val: str | None) -> float | None:
    """Convert a string to float, returning None for empty/invalid values."""
    if val is None or val == "" or val == ".." or val == "..":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_int(val: str | None) -> int | None:
    """Convert a string to int, returning None for empty/invalid values."""
    f = safe_float(val)
    if f is None:
        return None
    return int(f)


def ingest_census_profiles(db: sqlite3.Connection, skip_download: bool = False) -> int:
    """Download and ingest ABS Census CED demographic profiles.

    Returns the number of electorates ingested.
    """
    print("[abs_profiles] Ingesting ABS 2021 Census electorate profiles...")

    if not skip_download:
        download_datapack()

    if not ZIP_FILE.exists():
        raise FileNotFoundError(f"Census datapack zip not found at {ZIP_FILE}")

    zf = zipfile.ZipFile(ZIP_FILE, "r")

    # Read geography mapping (CED code -> electorate name)
    print("  Reading geography mapping...")
    geo_map = read_geography_mapping(zf)
    print(f"  Found {len(geo_map)} CED -> name mappings")

    # Read the key data tables
    print("  Reading G01 (population, indigenous, country of birth)...")
    g01 = {r["CED_CODE_2021"]: r for r in read_csv_from_zip(zf, "2021Census_G01_AUST_CED.csv")}

    print("  Reading G02 (medians and averages)...")
    g02 = {r["CED_CODE_2021"]: r for r in read_csv_from_zip(zf, "2021Census_G02_AUST_CED.csv")}

    print("  Reading G37 (housing tenure)...")
    g37 = {r["CED_CODE_2021"]: r for r in read_csv_from_zip(zf, "2021Census_G37_AUST_CED.csv")}

    print("  Reading G43 (labour force, education)...")
    g43 = {r["CED_CODE_2021"]: r for r in read_csv_from_zip(zf, "2021Census_G43_AUST_CED.csv")}

    zf.close()

    # Build the combined demographic records
    all_codes = sorted(set(g01.keys()) & set(g02.keys()) & set(g37.keys()) & set(g43.keys()))
    # Filter out "Total" rows, non-CED codes, and special geographic areas
    # CED codes ending in 94 = "No usual address", 97 = "Migratory/Offshore/Shipping"
    all_codes = [
        c for c in all_codes
        if c.startswith("CED")
        and not c.endswith("94")
        and not c.endswith("97")
    ]

    print(f"  Processing {len(all_codes)} electoral divisions...")

    records = []
    for code in all_codes:
        r01 = g01[code]
        r02 = g02[code]
        r37 = g37[code]
        r43 = g43[code]

        # Electorate name from geography mapping
        electorate_name = geo_map.get(code, code.replace("CED", "Electorate "))

        # Skip non-electorate entries that slipped through
        if any(skip in electorate_name for skip in ("Migratory", "No usual address", "Offshore")):
            continue

        # State from code
        state = get_state_from_code(code)

        # Population (total persons)
        population = safe_int(r01.get("Tot_P_P"))

        # Median income (weekly personal income)
        median_income = safe_float(r02.get("Median_tot_prsnl_inc_weekly"))

        # Median age
        median_age = safe_float(r02.get("Median_age_persons"))

        # Unemployment rate (direct percentage from G43)
        unemployment_rate = safe_float(r43.get("Percent_Unem_loyment_P"))

        # Housing tenure percentages
        owned_outright = safe_float(r37.get("O_OR_Total"))
        owned_mortgage = safe_float(r37.get("O_MTG_Total"))
        rented_total = safe_float(r37.get("R_Tot_Total"))
        total_dwellings = safe_float(r37.get("Total_Total"))

        homeownership_pct = None
        rental_pct = None
        if total_dwellings and total_dwellings > 0:
            if owned_outright is not None and owned_mortgage is not None:
                homeownership_pct = round(
                    ((owned_outright + owned_mortgage) / total_dwellings) * 100, 1
                )
            if rented_total is not None:
                rental_pct = round((rented_total / total_dwellings) * 100, 1)

        # Born overseas percentage
        born_elsewhere = safe_float(r01.get("Birthplace_Elsewhere_P"))
        born_australia = safe_float(r01.get("Birthplace_Australia_P"))
        born_overseas_pct = None
        total_birthplace = (born_elsewhere or 0) + (born_australia or 0)
        if total_birthplace > 0 and born_elsewhere is not None:
            born_overseas_pct = round((born_elsewhere / total_birthplace) * 100, 1)

        # University degree percentage (bachelor + postgrad + grad dip/cert)
        postgrad = safe_float(r43.get("non_sch_qual_PostGrad_Dgre_P")) or 0
        grad_dip = safe_float(r43.get("non_sch_qual_Gr_Dip_Gr_Crt_P")) or 0
        bachelor = safe_float(r43.get("non_sch_qual_Bchelr_Degree_P")) or 0
        persons_15plus = safe_float(r43.get("P_15_yrs_over_P"))
        university_pct = None
        if persons_15plus and persons_15plus > 0:
            university_pct = round(
                ((postgrad + grad_dip + bachelor) / persons_15plus) * 100, 1
            )

        # Indigenous percentage
        indigenous_total = safe_float(r01.get("Indigenous_P_Tot_P"))
        indigenous_pct = None
        if population and population > 0 and indigenous_total is not None:
            indigenous_pct = round((indigenous_total / population) * 100, 1)

        # Additional fields from G02
        median_rent_weekly = safe_float(r02.get("Median_rent_weekly"))
        median_mortgage_monthly = safe_float(r02.get("Median_mortgage_repay_monthly"))
        median_household_income_weekly = safe_float(r02.get("Median_tot_hhd_inc_weekly"))
        average_household_size = safe_float(r02.get("Average_household_size"))

        # Labour force participation
        labour_force_participation = safe_float(r43.get("Percnt_LabForc_prticipation_P"))

        records.append({
            "electorate_name": electorate_name,
            "state": state,
            "population": population,
            "median_income": median_income,
            "median_age": median_age,
            "unemployment_rate": unemployment_rate,
            "homeownership_pct": homeownership_pct,
            "rental_pct": rental_pct,
            "born_overseas_pct": born_overseas_pct,
            "university_pct": university_pct,
            "indigenous_pct": indigenous_pct,
            "median_rent_weekly": median_rent_weekly,
            "median_mortgage_monthly": median_mortgage_monthly,
            "median_household_income_weekly": median_household_income_weekly,
            "average_household_size": average_household_size,
            "labour_force_participation": labour_force_participation,
            "year": 2021,
            "source": "abs_census_2021",
        })

    # Insert into database
    print(f"  Inserting {len(records)} electorate demographic profiles...")
    inserted = 0
    for rec in records:
        try:
            db.execute(
                """
                INSERT OR REPLACE INTO electorate_demographics (
                    electorate_name, state, population, median_income, median_age,
                    unemployment_rate, homeownership_pct, rental_pct, born_overseas_pct,
                    university_pct, indigenous_pct, median_rent_weekly,
                    median_mortgage_monthly, median_household_income_weekly,
                    average_household_size, labour_force_participation, year, source
                ) VALUES (
                    :electorate_name, :state, :population, :median_income, :median_age,
                    :unemployment_rate, :homeownership_pct, :rental_pct, :born_overseas_pct,
                    :university_pct, :indigenous_pct, :median_rent_weekly,
                    :median_mortgage_monthly, :median_household_income_weekly,
                    :average_household_size, :labour_force_participation, :year, :source
                )
                """,
                rec,
            )
            inserted += 1
        except sqlite3.IntegrityError as e:
            print(f"  Warning: skipping {rec['electorate_name']}: {e}")

    db.commit()
    print(f"  Inserted {inserted} electorate profiles")

    # Print summary stats
    _print_summary(db)

    return inserted


def _print_summary(db: sqlite3.Connection) -> None:
    """Print summary statistics for the ingested data."""
    row = db.execute(
        """
        SELECT COUNT(*) as count,
               ROUND(AVG(median_income), 0) as avg_income,
               ROUND(MIN(median_income), 0) as min_income,
               ROUND(MAX(median_income), 0) as max_income,
               ROUND(AVG(unemployment_rate), 1) as avg_unemp,
               ROUND(AVG(university_pct), 1) as avg_uni,
               ROUND(AVG(indigenous_pct), 1) as avg_indigenous,
               ROUND(AVG(born_overseas_pct), 1) as avg_overseas
        FROM electorate_demographics
        """
    ).fetchone()

    print(f"\n  Summary ({row['count']} electorates):")
    print(f"    Median weekly income: ${row['min_income']}-${row['max_income']} (avg ${row['avg_income']})")
    print(f"    Avg unemployment: {row['avg_unemp']}%")
    print(f"    Avg university educated: {row['avg_uni']}%")
    print(f"    Avg indigenous: {row['avg_indigenous']}%")
    print(f"    Avg born overseas: {row['avg_overseas']}%")

    # Top 5 richest and poorest
    print("\n  Highest median income:")
    for r in db.execute(
        "SELECT electorate_name, median_income, state FROM electorate_demographics ORDER BY median_income DESC LIMIT 5"
    ).fetchall():
        print(f"    {r['electorate_name']} ({r['state']}): ${r['median_income']}/week")

    print("\n  Lowest median income:")
    for r in db.execute(
        "SELECT electorate_name, median_income, state FROM electorate_demographics WHERE median_income IS NOT NULL ORDER BY median_income ASC LIMIT 5"
    ).fetchall():
        print(f"    {r['electorate_name']} ({r['state']}): ${r['median_income']}/week")

    print("\n  Highest unemployment:")
    for r in db.execute(
        "SELECT electorate_name, unemployment_rate, state FROM electorate_demographics ORDER BY unemployment_rate DESC LIMIT 5"
    ).fetchall():
        print(f"    {r['electorate_name']} ({r['state']}): {r['unemployment_rate']}%")


def analyze_representation_gaps(db: sqlite3.Connection) -> dict:
    """Cross-reference electorate demographics with MP voting records.

    Identifies cases where MPs may not represent their constituents' interests:
    - Low-income electorates where MPs voted against cost-of-living measures
    - High-unemployment electorates where MPs voted against jobs bills
    - High-rental electorates where MPs voted against housing measures
    - High-indigenous electorates where MPs voted against indigenous affairs

    Returns a dict of stories suitable for the analysis_cache.
    """
    print("\n[abs_profiles] Analyzing representation gaps...")

    stories = []

    # --- Cost of living: low-income electorates ---
    # Find divisions related to cost of living
    col_division_ids = db.execute(
        """
        SELECT division_id, name, date FROM divisions
        WHERE name LIKE '%cost of living%'
           OR name LIKE '%energy price%'
           OR name LIKE '%grocery%'
           OR name LIKE '%housing afford%'
           OR name LIKE '%rental%'
           OR name LIKE '%wage%'
           OR name LIKE '%minimum wage%'
           OR name LIKE '%household%'
        ORDER BY date DESC
        """
    ).fetchall()

    if col_division_ids:
        print(f"  Found {len(col_division_ids)} cost-of-living related divisions")

        # MPs in the lowest-income quartile
        low_income_threshold = db.execute(
            """
            SELECT median_income FROM electorate_demographics
            WHERE median_income IS NOT NULL
            ORDER BY median_income ASC
            LIMIT 1 OFFSET (SELECT COUNT(*)/4 FROM electorate_demographics WHERE median_income IS NOT NULL)
            """
        ).fetchone()

        if low_income_threshold:
            threshold = low_income_threshold["median_income"]
            print(f"  Low-income threshold (Q1): ${threshold}/week")

            for div in col_division_ids:
                # Find MPs from low-income electorates who voted 'no' on these measures
                no_voters = db.execute(
                    """
                    SELECT m.full_name, m.party, m.electorate, ed.median_income,
                           ed.unemployment_rate, v.vote
                    FROM votes v
                    JOIN members m ON v.person_id = m.person_id
                    JOIN electorate_demographics ed ON ed.electorate_name = m.electorate
                    WHERE v.division_id = ?
                      AND v.vote = 'no'
                      AND ed.median_income <= ?
                      AND m.chamber = 'representatives'
                    ORDER BY ed.median_income ASC
                    """,
                    (div["division_id"], threshold),
                ).fetchall()

                if no_voters:
                    stories.append({
                        "type": "cost_of_living_disconnect",
                        "division": div["name"],
                        "division_date": div["date"],
                        "description": (
                            f"{len(no_voters)} MP(s) from low-income electorates "
                            f"voted against: {div['name']}"
                        ),
                        "mps": [
                            {
                                "name": r["full_name"],
                                "party": r["party"],
                                "electorate": r["electorate"],
                                "median_income": r["median_income"],
                                "unemployment": r["unemployment_rate"],
                            }
                            for r in no_voters
                        ],
                    })

    # --- Housing: high-rental electorates ---
    housing_division_ids = db.execute(
        """
        SELECT division_id, name, date FROM divisions
        WHERE name LIKE '%housing%'
           OR name LIKE '%rent%'
           OR name LIKE '%tenant%'
           OR name LIKE '%homelessness%'
           OR name LIKE '%first home%'
        ORDER BY date DESC
        """
    ).fetchall()

    if housing_division_ids:
        print(f"  Found {len(housing_division_ids)} housing-related divisions")

        high_rental_threshold = db.execute(
            """
            SELECT rental_pct FROM electorate_demographics
            WHERE rental_pct IS NOT NULL
            ORDER BY rental_pct DESC
            LIMIT 1 OFFSET (SELECT COUNT(*)/4 FROM electorate_demographics WHERE rental_pct IS NOT NULL)
            """
        ).fetchone()

        if high_rental_threshold:
            threshold = high_rental_threshold["rental_pct"]
            for div in housing_division_ids:
                no_voters = db.execute(
                    """
                    SELECT m.full_name, m.party, m.electorate, ed.rental_pct,
                           ed.median_rent_weekly, v.vote
                    FROM votes v
                    JOIN members m ON v.person_id = m.person_id
                    JOIN electorate_demographics ed ON ed.electorate_name = m.electorate
                    WHERE v.division_id = ?
                      AND v.vote = 'no'
                      AND ed.rental_pct >= ?
                      AND m.chamber = 'representatives'
                    ORDER BY ed.rental_pct DESC
                    """,
                    (div["division_id"], threshold),
                ).fetchall()

                if no_voters:
                    stories.append({
                        "type": "housing_disconnect",
                        "division": div["name"],
                        "division_date": div["date"],
                        "description": (
                            f"{len(no_voters)} MP(s) from high-rental electorates "
                            f"voted against: {div['name']}"
                        ),
                        "mps": [
                            {
                                "name": r["full_name"],
                                "party": r["party"],
                                "electorate": r["electorate"],
                                "rental_pct": r["rental_pct"],
                                "median_rent": r["median_rent_weekly"],
                            }
                            for r in no_voters
                        ],
                    })

    # --- Indigenous affairs: high-indigenous electorates ---
    indigenous_division_ids = db.execute(
        """
        SELECT division_id, name, date FROM divisions
        WHERE name LIKE '%indigenous%'
           OR name LIKE '%aboriginal%'
           OR name LIKE '%first nations%'
           OR name LIKE '%closing the gap%'
           OR name LIKE '%native title%'
           OR name LIKE '%voice%referendum%'
           OR name LIKE '%reconciliation%'
        ORDER BY date DESC
        """
    ).fetchall()

    if indigenous_division_ids:
        print(f"  Found {len(indigenous_division_ids)} indigenous-affairs related divisions")

        high_indigenous_threshold = db.execute(
            """
            SELECT indigenous_pct FROM electorate_demographics
            WHERE indigenous_pct IS NOT NULL
            ORDER BY indigenous_pct DESC
            LIMIT 1 OFFSET (SELECT COUNT(*)/4 FROM electorate_demographics WHERE indigenous_pct IS NOT NULL)
            """
        ).fetchone()

        if high_indigenous_threshold:
            threshold = high_indigenous_threshold["indigenous_pct"]
            for div in indigenous_division_ids:
                no_voters = db.execute(
                    """
                    SELECT m.full_name, m.party, m.electorate, ed.indigenous_pct, v.vote
                    FROM votes v
                    JOIN members m ON v.person_id = m.person_id
                    JOIN electorate_demographics ed ON ed.electorate_name = m.electorate
                    WHERE v.division_id = ?
                      AND v.vote = 'no'
                      AND ed.indigenous_pct >= ?
                      AND m.chamber = 'representatives'
                    ORDER BY ed.indigenous_pct DESC
                    """,
                    (div["division_id"], threshold),
                ).fetchall()

                if no_voters:
                    stories.append({
                        "type": "indigenous_disconnect",
                        "division": div["name"],
                        "division_date": div["date"],
                        "description": (
                            f"{len(no_voters)} MP(s) from high-indigenous electorates "
                            f"voted against: {div['name']}"
                        ),
                        "mps": [
                            {
                                "name": r["full_name"],
                                "party": r["party"],
                                "electorate": r["electorate"],
                                "indigenous_pct": r["indigenous_pct"],
                            }
                            for r in no_voters
                        ],
                    })

    # Store in analysis_cache
    result = {
        "total_stories": len(stories),
        "by_type": {},
        "stories": stories,
    }
    for s in stories:
        t = s["type"]
        result["by_type"][t] = result["by_type"].get(t, 0) + 1

    db.execute(
        "INSERT OR REPLACE INTO analysis_cache (key, value, updated_at) VALUES (?, ?, datetime('now'))",
        ("stories_unrepresentative", json.dumps(result)),
    )
    db.commit()

    print(f"\n  Representation gap analysis complete:")
    print(f"    Total stories: {len(stories)}")
    for t, count in result["by_type"].items():
        print(f"    {t}: {count}")

    return result


def main():
    parser = argparse.ArgumentParser(
        description="Ingest ABS Census electorate demographic profiles"
    )
    parser.add_argument(
        "--no-download",
        action="store_true",
        help="Skip download, use cached zip file",
    )
    parser.add_argument(
        "--analyze",
        action="store_true",
        help="Run cross-reference analysis (representation gaps)",
    )
    parser.add_argument(
        "--analyze-only",
        action="store_true",
        help="Skip ingest, only run analysis",
    )
    args = parser.parse_args()

    db = get_db()
    db.execute("PRAGMA busy_timeout = 600000")
    init_db(db)

    if not args.analyze_only:
        count = ingest_census_profiles(db, skip_download=args.no_download)
        print(f"\n[abs_profiles] Done. Ingested {count} electorate profiles.")

    if args.analyze or args.analyze_only:
        analyze_representation_gaps(db)


if __name__ == "__main__":
    main()
